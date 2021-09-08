from openpyxl import load_workbook
from openpyxl import Workbook


class ListDict_txt_xlsx:

     def __init__(self, list_):

         self.list = list_
         self.dict_ = None

     def convert_list_to_dict(self):

          _dict = {}

          for obj in self.list:
               _dict[obj] = obj

          self.dict_ = _dict

          return _dict

     def save_in_file(self):

          with open('converted_dict.txt', 'w') as file_dict:
               file_dict.write(str(self.dict_))
               file_dict.close()

     def read_from_file(self):
          
          with open('converted_dict.txt', 'r') as file:
               return eval(file.read())
               
               

     def save_in_xlsx(self, key, value):

          wb = Workbook()
          sheet = wb.active
          sheet['A1'] = 'Key'
          sheet['B1'] = 'Value'

          for row, (key, value) in enumerate(self.dict_.items(), start=2):
               sheet[f'A{row}'] = key
               sheet[f'B{row}'] = value

          wb.save('dict.xlsx')
          wb.close()

     def read_in_xlsx(self):

          _dict = {}

          wb = load_workbook('dict.xlsx')
          sheet = wb.active
          
          for key, value in sheet.iter_rows(min_row=2, max_col=2, max_row=6, values_only=True):
              _dict[key] = key
              _dict[value] = value              

          return _dict
          

object_list = ListDict_txt_xlsx(list_ = [1, 2, 3, 4, 5])

convert = object_list.convert_list_to_dict()

save_file_txt = object_list.save_in_file()

read_file_txt = object_list.read_from_file()
print(f'Reading a dictionary from a text file {read_file_txt}')

save_file_xlsx = object_list.save_in_xlsx('key', 'value')

read_file_xlsx = object_list.read_in_xlsx()
print(f'Reading a dictionary from Exel file {read_file_xlsx}')
