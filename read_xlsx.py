# # Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook


wb = load_workbook('hello.xlsx')

# Get sheet names
print(wb.sheetnames)
sheet_name = wb.sheetnames[0]

sheet = wb[sheet_name]

# print(sheet['A1'].value)
row = len(list(sheet.values))
print(row)
column = len(list(sheet.values)[0])


for _row in range(1, row):
    for _column in range(1, column):

        value = sheet.cell(row=_row, column=_column).value
        print(value)


# Write
sheet.cell(row=12, column=12).value = "oo"
wb.save('hello.xlsx')
